#!/usr/bin/env python3
"""
Invoice PDF Extractor — Dream Smiles Dental
============================================
Handles all known lab invoice formats:
  - 3 Dental       : Summary invoice, table of OrderID/Patient/Total
  - Dent8          : Statement, table of Invoice/Patient/Amount/Payments/Balance
  - Innovate Dental: Statement, table of Invoice/Patient/Amount/Payments/Balance
  - S4S            : Statement, table of Advice/Patient/Total
  - Hall Dental    : Statement, table of Job/Invoice/Patient/OrigAmt/Balance
  - Carl Kearney   : Statement, list of invoices with patient names
  - Aesthetic World: Statement, single line invoice per patient
  - Digital Pros.  : Monthly statement, invoices + opening balance + payments

Usage:
    python extract_invoice.py <invoice.pdf>
    python extract_invoice.py <folder>
"""

import sys
import os
import re
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    os.system("python -m pip install pdfplumber")
    import pdfplumber

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    os.system("python -m pip install openpyxl")
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

TRACKER_PATH = Path(__file__).parent / "Invoice Tracker 2026.xlsx"

KNOWN_LABS = [
    "Hall Dental Studio", "Innovate Dental", "Dent8", "Invisalign",
    "Carl Kearney", "Digital Prosthetics", "S4S", "Aesthetic World",
    "3 Dental", "Avant Garde", "Boutique Whitening"
]

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def extract_pdf(pdf_path):
    pages_text, all_tables = [], []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            pages_text.append(text)
            all_tables.extend(page.extract_tables() or [])
    return "\n".join(pages_text), all_tables

def clean_amount(raw):
    if raw is None:
        return None
    cleaned = re.sub(r"[£$,\s]", "", str(raw).strip())
    try:
        val = float(cleaned)
        return round(val, 2) if val != 0 else None
    except ValueError:
        return None

def detect_lab(text):
    # Some PDFs have stray U+0000 inside words (e.g. "Bou\x00que Whitening" —
    # the "ti" ligature mis-decodes as a single null byte). Strip control chars
    # before matching so the substring search is robust.
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    lowered = cleaned.lower()
    for lab in KNOWN_LABS:
        if lab.lower() in lowered:
            return lab
    # Per-lab fallback heuristics for PDFs whose name is mangled in the text
    # layer. Each anchor must be specific enough to identify only that lab.
    if 'BW' in cleaned and ('whitening' in lowered or 'thermoform' in lowered):
        return 'Boutique Whitening'
    # Detect by invoice number prefix
    if re.search(r'INV-D[0-9]', cleaned):
        return "Dent8"
    if re.search(r'INV-IN[0-9]', cleaned):
        return "Innovate Dental"
    return None

def find_next_row(ws):
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row

def clean_str(v):
    if isinstance(v, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', v).strip()
    return v

def get_existing_keys(ws):
    """Return a set of (invoice_number, invoice_date) tuples already in the tracker."""
    keys = set()
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        date = str(ws.cell(row=row, column=1).value or "").strip().lower()
        inv  = str(ws.cell(row=row, column=2).value or "").strip().lower()
        if inv and inv not in ("invoice number", "statement total", "summary total"):
            keys.add((inv, date))
        row += 1
    return keys

def append_rows(rows):
    if not TRACKER_PATH.exists():
        print(f"  ERROR: Tracker not found at {TRACKER_PATH}")
        return 0
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Invoice Tracker"]
    alt_fill = PatternFill("solid", start_color="EAF0FB")
    currency_fmt = '£#,##0.00;(£#,##0.00);"-"'
    next_row = find_next_row(ws)
    existing_keys = get_existing_keys(ws)

    written, skipped = 0, 0
    for data in rows:
        inv_no = clean_str(data.get("invoice_number") or "").lower()
        inv_date = clean_str(str(data.get("invoice_date") or "")).lower()
        key = (inv_no, inv_date)

        # Skip duplicates (ignore summary/total rows from duplicate check)
        if inv_no and inv_no not in ("statement total", "summary total") and key in existing_keys:
            skipped += 1
            continue

        vals = [
            clean_str(data.get("invoice_date")), clean_str(data.get("invoice_number")),
            clean_str(data.get("job_reference")), clean_str(data.get("patient_name")),
            clean_str(data.get("laboratory_name")), data.get("invoiced_amount"),
            data.get("payments_made"), data.get("balance"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            if col >= 6:
                cell.number_format = currency_fmt
            if next_row % 2 == 0:
                cell.fill = alt_fill
        existing_keys.add(key)
        next_row += 1
        written += 1

    wb.save(TRACKER_PATH)
    if skipped:
        print(f"  ⚠ {skipped} duplicate(s) skipped")
    return written

def deduplicate():
    """Remove duplicate rows from the tracker based on invoice number + date."""
    if not TRACKER_PATH.exists():
        print(f"ERROR: Tracker not found at {TRACKER_PATH}")
        return
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Invoice Tracker"]
    alt_fill = PatternFill("solid", start_color="EAF0FB")
    currency_fmt = '£#,##0.00;(£#,##0.00);"-"'

    # Collect all data rows
    data_rows = []
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        data_rows.append([ws.cell(row=row, column=col).value for col in range(1, 9)])
        row += 1

    total_before = len(data_rows)
    seen = set()
    unique_rows = []
    removed = 0

    for r in data_rows:
        inv_date = str(r[0] or "").strip().lower()
        inv_no   = str(r[1] or "").strip().lower()
        key = (inv_no, inv_date)
        # Always keep summary/total rows
        if inv_no in ("statement total", "summary total", ""):
            unique_rows.append(r)
        elif key in seen:
            removed += 1
        else:
            seen.add(key)
            unique_rows.append(r)

    # Clear existing data rows
    for r in range(3, 3 + total_before + 1):
        for col in range(1, 9):
            ws.cell(row=r, column=col).value = None

    # Re-write unique rows
    for i, r in enumerate(unique_rows):
        write_row = 3 + i
        for col, val in enumerate(r, 1):
            cell = ws.cell(row=write_row, column=col, value=val)
            if col >= 6:
                cell.number_format = currency_fmt
            if write_row % 2 == 0:
                cell.fill = alt_fill

    wb.save(TRACKER_PATH)
    print(f"\nDeduplication complete:")
    print(f"  Rows before : {total_before}")
    print(f"  Duplicates removed: {removed}")
    print(f"  Rows after  : {len(unique_rows)}")

# ─────────────────────────────────────────────
# Format detectors
# ─────────────────────────────────────────────

def detect_format(text):
    t = text.lower()
    if re.search(r'summary\s*no', t) or re.search(r'orderid\s+patient', t):
        return "3dental"
    if "hall dental" in t or "halldentalstudio" in t:
        return "hall"
    if "carl kearney" in t:
        return "carlkearney"
    if "aesthetic world" in t:
        return "aestheticworld"
    if "digital prosthetics" in t:
        return "digitalprothetics"
    if "s4s" in t and "advice" in t:
        return "s4s"
    # Dent8 and Innovate share the same portal format (INV-D or INV-IN prefix)
    if re.search(r'inv-[a-z0-9]+', t) and ("invoice amount" in t or "pa\x00ent" in t or "pa ent" in t):
        return "dent8_innovate"
    return "standard"

# ─────────────────────────────────────────────
# Parsers
# ─────────────────────────────────────────────

def parse_3dental(text, tables, lab):
    inv_no = None
    m = re.search(r'summary\s*no[.\s:]*([A-Z0-9]+)', text, re.IGNORECASE)
    if m:
        inv_no = m.group(1).strip()

    invoice_date = None
    for pat in [r'PAYMENT\s+DUE\s+(\d{4}-\d{2}-\d{2})', r'TO\s+(\d{4}-\d{2}-\d{2})', r'FROM\s+(\d{4}-\d{2}-\d{2})']:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            invoice_date = m.group(1)
            break

    balance = None
    m = re.search(r'BALANCE\s+DUE\s+[£$]?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        balance = clean_amount(m.group(1))

    total = None
    m = re.search(r'\bTOTAL\b\s+[£$]?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        total = clean_amount(m.group(1))

    seen_orders = {}
    for table in tables:
        if not table:
            continue
        header = [str(c).lower().strip() if c else "" for c in table[0]]
        if "orderid" not in " ".join(header):
            continue
        col = {}
        for i, h in enumerate(header):
            if "order" in h: col["order"] = i
            elif "patient" in h: col["patient"] = i
            elif "date" in h: col["date"] = i
            elif "total" in h: col["total"] = i

        for row in table[1:]:
            if not row or all(c is None for c in row):
                continue
            oid = str(row[col.get("order", 0)] or "").strip()
            patient = str(row[col.get("patient", 1)] or "").strip()
            amount = clean_amount(row[col.get("total", -1)] if col.get("total") is not None else None)
            if not oid or oid.lower() in ("orderid", ""):
                continue
            if oid not in seen_orders:
                seen_orders[oid] = {"patient": patient, "total": 0.0}
            if amount:
                seen_orders[oid]["total"] = round(seen_orders[oid]["total"] + amount, 2)

    rows = []
    for oid, info in seen_orders.items():
        rows.append({
            "invoice_date": invoice_date, "invoice_number": inv_no,
            "job_reference": oid, "patient_name": info["patient"].title(),
            "laboratory_name": lab, "invoiced_amount": info["total"],
            "payments_made": None, "balance": None,
        })
    rows.append({
        "invoice_date": invoice_date, "invoice_number": inv_no,
        "job_reference": "SUMMARY TOTAL", "patient_name": f"({len(seen_orders)} orders)",
        "laboratory_name": lab, "invoiced_amount": total or balance,
        "payments_made": None, "balance": balance,
    })
    return rows

def parse_dent8_innovate(text, tables, lab):
    """
    Handles Dent8 and Innovate Dental portal statements.
    These PDFs merge all rows into a single table cell, so we parse from raw text.
    Pattern per line: date  INV-XXXXX  Patient Name  due_date  £amount  £payments  £balance
    """
    stmt_date = None
    m = re.search(r'Date\s+(\d{2}/\d{2}/\d{4})', text)
    if m:
        stmt_date = m.group(1)

    total_due = None
    m = re.search(r'Total\s+Amount\s+Due\s+[£$]?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        total_due = clean_amount(m.group(1))

    # Detect lab from known patterns in text or invoice number prefix
    if not lab:
        if "dent8" in text.lower() or "blackpool" in text.lower() or "garton" in text.lower():
            lab = "Dent8"
        elif "innovate" in text.lower() or "oldham" in text.lower() or "langdale" in text.lower():
            lab = "Innovate Dental"
        elif re.search(r'INV-D\d+', text):
            lab = "Dent8"
        elif re.search(r'INV-IN\d+', text):
            lab = "Innovate Dental"

    rows = []
    # Each invoice line: date  INV-XXXXXX-X-X  Patient Name  due_date  £amt  £pay  £bal
    inv_pattern = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+(INV-[A-Z0-9\-]+)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})\s+[£$]([\d,]+\.\d{2})\s+[£$]([\d,]+\.\d{2})\s+[£$]([\d,]+\.\d{2})'
    )
    for m in inv_pattern.finditer(text):
        date, inv_no, patient, due_date, amount, payments, balance = m.groups()
        rows.append({
            "invoice_date": date, "invoice_number": inv_no,
            "job_reference": None, "patient_name": patient.strip().title(),
            "laboratory_name": lab, "invoiced_amount": clean_amount(amount),
            "payments_made": clean_amount(payments),
            "balance": clean_amount(balance),
        })

    if rows and total_due:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": "STATEMENT TOTAL",
            "job_reference": None, "patient_name": f"({len(rows)} invoices)",
            "laboratory_name": lab, "invoiced_amount": None,
            "payments_made": None, "balance": total_due,
        })
    return rows

def parse_hall(text, tables, lab):
    """Hall Dental Studio statement format."""
    stmt_date = None
    m = re.search(r'Unpaid items up to[:\s]+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if m:
        stmt_date = m.group(1)

    total_due = None
    m = re.search(r'AMOUNT\s+DUE[:\s]+[£$]?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        total_due = clean_amount(m.group(1))

    rows = []
    # Hall has used two slightly different statement layouts; we try both.
    #
    # Old layout (with balance column):
    #   24/02/26  T61947  -  INV68272  U BUKSH  3900.00  3078.50
    #
    # Newer layout (no balance, optional B/R/L code suffix on the patient):
    #   12/01/26  INV67806  T61617  H BRAY  B1  150.00
    #   22/01/26  INV67923  T61476  K BATH R 64321 12346 L  2925.00
    OLD_PATTERN = re.compile(
        r'(\d{2}/\d{2}/\d{2})\s+(T\d+)\s+-\s+(INV\d+)\s+([A-Z][A-Z\s]+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})'
    )
    NEW_PATTERN = re.compile(
        # Date  INV...  T...  PATIENT NAME (caps + spaces, allow trailing tokens like "B1", "R 4 L")  AMOUNT
        # The patient capture is non-greedy and stops at the last numeric amount on the line.
        # Don't anchor on $ — pdfplumber may merge follow-on description text onto the same line.
        r'(\d{2}/\d{2}/\d{2})\s+(INV\d+)\s+(T\d+)\s+([A-Z][A-Z0-9\s]+?)\s+([\d,]+\.\d{2})(?:\s|$)',
        re.MULTILINE,
    )
    seen_invoices = set()
    for m in OLD_PATTERN.finditer(text):
        date, job, inv, patient, orig, balance = m.groups()
        seen_invoices.add(inv)
        rows.append({
            "invoice_date": date, "invoice_number": inv,
            "job_reference": job, "patient_name": patient.strip().title(),
            "laboratory_name": lab, "invoiced_amount": clean_amount(orig),
            "payments_made": None, "balance": clean_amount(balance),
        })
    for m in NEW_PATTERN.finditer(text):
        date, inv, job, patient, orig = m.groups()
        if inv in seen_invoices:
            continue
        rows.append({
            "invoice_date": date, "invoice_number": inv,
            "job_reference": job, "patient_name": patient.strip().title(),
            "laboratory_name": lab, "invoiced_amount": clean_amount(orig),
            "payments_made": None, "balance": None,
        })

    if rows:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": "STATEMENT TOTAL",
            "job_reference": None, "patient_name": f"({len(rows)} invoices)",
            "laboratory_name": lab, "invoiced_amount": None,
            "payments_made": None, "balance": total_due,
        })
    return rows

def parse_carlkearney(text, tables, lab):
    """
    Carl Kearney statement.
    Format: date  Invoice No.IXXXXXX: Due date.  amount  open_amount
            (next line) PATIENT NAME
    """
    stmt_no = None
    m = re.search(r'STATEMENT\s+NO[.\s:]*(\d+)', text, re.IGNORECASE)
    if m:
        stmt_no = m.group(1)

    stmt_date = None
    m = re.search(r'DATE\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if m:
        stmt_date = m.group(1)

    total_due = None
    m = re.search(r'TOTAL\s+DUE\s+[£$]([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        total_due = clean_amount(m.group(1))

    rows = []
    # Two-line pattern: invoice line followed by patient name line
    # e.g.: "12/02/2026 Invoice No.I001694: Due 1,940.00 1,940.00\n28/02/2026. ANDY TARBURTON"
    inv_pattern = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+Invoice\s+No\.(I\d+)[^\n]+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*\n'
        r'[\d/\.]*\s*([A-Z][A-Z\s]+?)(?:\n|$)',
        re.IGNORECASE
    )
    for m in inv_pattern.finditer(text):
        date, inv_no, amount, open_amt, patient = m.groups()
        rows.append({
            "invoice_date": date, "invoice_number": inv_no,
            "job_reference": f"STMT-{stmt_no}" if stmt_no else None,
            "patient_name": patient.strip().title(),
            "laboratory_name": lab, "invoiced_amount": clean_amount(amount),
            "payments_made": None, "balance": clean_amount(open_amt),
        })

    if rows:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": f"STMT-{stmt_no}",
            "job_reference": None, "patient_name": f"({len(rows)} invoices)",
            "laboratory_name": lab, "invoiced_amount": None,
            "payments_made": None, "balance": total_due,
        })
    return rows

def parse_aestheticworld(text, tables, lab):
    """
    Aesthetic World statement.
    Line format: date  INV-AW30282-W-\n1  PATIENT NAME  description  net  vat  gross
    The invoice number sometimes wraps to the next line.
    """
    stmt_date = None
    m = re.search(r'Date\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if m:
        stmt_date = m.group(1)

    total_due = None
    m = re.search(r'Total\s+Due\s+[£$\s]*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        total_due = clean_amount(m.group(1))

    # Normalise line-wrapped invoice numbers like "INV-AW30282-W-\n1" -> "INV-AW30282-W-1"
    normalised = re.sub(r'(INV-[\w\-]+)-\s*\n(\d+)', r'\1-\2', text)

    rows = []
    # Pattern: date  INV-xxxxx  PATIENT NAME  description...  net  vat  gross
    inv_pattern = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+(INV-[\w\-]+)\s+([A-Z][A-Z\s]{2,40}?)\s{2,}.+?\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})',
        re.IGNORECASE
    )
    for m in inv_pattern.finditer(normalised):
        date, inv_no, patient, net, vat, gross = m.groups()
        rows.append({
            "invoice_date": date, "invoice_number": inv_no.strip(),
            "job_reference": None, "patient_name": patient.strip().title(),
            "laboratory_name": lab, "invoiced_amount": clean_amount(gross),
            "payments_made": None, "balance": clean_amount(gross),
        })

    if not rows:
        # Fallback: simpler pattern just grabbing date, inv, last 3 numbers
        simple = re.compile(r'(\d{2}/\d{2}/\d{4})\s+(INV-[\w\-]+)\s+(.+?)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s*$', re.MULTILINE)
        for m in simple.finditer(normalised):
            date, inv_no, rest, net, vat, gross = m.groups()
            # Patient is first ALL-CAPS words in rest
            patient_m = re.match(r'([A-Z][A-Z\s]+?)(?:\s{2,}|\d)', rest)
            patient = patient_m.group(1).strip().title() if patient_m else rest.strip().title()
            rows.append({
                "invoice_date": date, "invoice_number": inv_no.strip(),
                "job_reference": None, "patient_name": patient,
                "laboratory_name": lab, "invoiced_amount": clean_amount(gross),
                "payments_made": None, "balance": clean_amount(gross),
            })

    if rows:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": "STATEMENT TOTAL",
            "job_reference": None, "patient_name": f"({len(rows)} invoices)",
            "laboratory_name": lab, "invoiced_amount": total_due,
            "payments_made": None, "balance": total_due,
        })
    return rows

def parse_digitalprothetics(text, tables, lab):
    """
    Digital Prosthetics monthly statement.
    Uses the extracted table directly — rows have: Date, Type, Provider, Patient, Amount, Balance
    Invoice rows have type like "Invoice INV44458"
    Payment rows have type "Payment" with negative amount.
    """
    stmt_date = None
    m = re.search(r'Statement\s+Date\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if m:
        stmt_date = m.group(1)

    total_due = None
    m = re.search(r'Total\s+Due\s+([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        total_due = clean_amount(m.group(1))

    payment_total = None
    rows = []

    for table in tables:
        if not table:
            continue
        header = [str(c or "").lower().strip() for c in table[0]]
        if "type" not in " ".join(header) and "invoice" not in " ".join(header):
            continue

        # Map columns
        col = {}
        for i, h in enumerate(header):
            if h == "date": col["date"] = i
            elif h == "type": col["type"] = i
            elif "provider" in h: col["provider"] = i
            elif "patient" in h: col["patient"] = i
            elif "amount" in h: col["amount"] = i
            elif "balance" in h: col["balance"] = i

        for row in table[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            type_val = str(row[col["type"]] or "").strip() if "type" in col else ""
            # Invoice row
            inv_m = re.match(r'Invoice\s+(INV\d+)', type_val, re.IGNORECASE)
            if inv_m:
                inv_no = inv_m.group(1)
                date = str(row[col["date"]] or stmt_date).strip() if "date" in col else stmt_date
                patient = str(row[col.get("patient", 3)] or "").strip().title()
                amount = clean_amount(row[col["amount"]]) if "amount" in col else None
                balance = clean_amount(row[col["balance"]]) if "balance" in col else None
                rows.append({
                    "invoice_date": date, "invoice_number": inv_no,
                    "job_reference": None, "patient_name": patient or None,
                    "laboratory_name": lab, "invoiced_amount": amount,
                    "payments_made": None, "balance": balance,
                })
            # Payment row
            elif "payment" in type_val.lower():
                amt = row[col["amount"]] if "amount" in col else None
                payment_total = clean_amount(str(amt).replace("-", "")) if amt else None

    if rows:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": "STATEMENT TOTAL",
            "job_reference": None, "patient_name": f"({len(rows)} invoices)",
            "laboratory_name": lab, "invoiced_amount": None,
            "payments_made": payment_total, "balance": total_due,
        })
    return rows

def parse_s4s(text, tables, lab):
    """
    S4S statement — uses the structured table which has columns:
    Date | Type | Method | Ref | Details | Credit | Original Total | Unapplied Credit | Amount Outstanding | Total
    Advice rows = invoices. Payment rows = payments made.
    """
    stmt_date = None
    m = re.search(r'Date[:\s]+(\d{2}\s+\w+\s+\d{4})', text, re.IGNORECASE)
    if m:
        stmt_date = m.group(1)

    ref_no = None
    m = re.search(r'Reference\s+No[:\s]+(\d+)', text, re.IGNORECASE)
    if m:
        ref_no = m.group(1)

    balance_due = None
    m = re.search(r'Balance\s+Due\s+[£$]?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
    if m:
        balance_due = clean_amount(m.group(1))

    payment_total = 0.0
    rows = []

    for table in tables:
        if not table or len(table) < 2:
            continue
        header = [str(c or "").lower().replace("\n", " ").strip() for c in table[0]]
        if "type" not in " ".join(header) or "details" not in " ".join(header):
            continue

        col = {}
        for i, h in enumerate(header):
            if h == "date": col["date"] = i
            elif h == "type": col["type"] = i
            elif h == "ref": col["ref"] = i
            elif h == "details": col["details"] = i
            elif "credit" in h and "unapplied" not in h and "original" not in h: col["credit"] = i
            elif "original" in h: col["orig"] = i

        for row in table[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            type_val = str(row[col.get("type", 1)] or "").strip().lower()
            details = str(row[col.get("details", 4)] or "").strip()
            date = str(row[col.get("date", 0)] or "").strip()
            ref = str(row[col.get("ref", 3)] or "").strip()
            orig = clean_amount(row[col.get("orig", 6)]) if "orig" in col else None
            credit = clean_amount(row[col.get("credit", 5)]) if "credit" in col else None

            if type_val == "advice":
                # Extract patient name from details: "Patient Name (CODE)"
                patient_m = re.match(r'(.+?)\s*\([\w\d]+\)', details)
                patient = patient_m.group(1).strip().title() if patient_m else details.title()
                rows.append({
                    "invoice_date": date or stmt_date, "invoice_number": ref,
                    "job_reference": ref_no, "patient_name": patient,
                    "laboratory_name": lab, "invoiced_amount": orig or credit,
                    "payments_made": None, "balance": None,
                })
            elif type_val == "payment" and credit:
                payment_total = round(payment_total + credit, 2)

    if rows:
        rows.append({
            "invoice_date": stmt_date, "invoice_number": "STATEMENT TOTAL",
            "job_reference": ref_no, "patient_name": f"({len(rows)} items)",
            "laboratory_name": lab, "invoiced_amount": None,
            "payments_made": payment_total if payment_total > 0 else None,
            "balance": balance_due,
        })
    return rows

def parse_standard(text, lab):
    data = {
        "invoice_date": None, "invoice_number": None, "job_reference": None,
        "patient_name": None, "laboratory_name": lab,
        "invoiced_amount": None, "payments_made": None, "balance": None,
    }
    for pat in [r"(?:invoice\s*date|date)[:\s]+(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
                r"\b(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})\b", r"\b(\d{4}-\d{2}-\d{2})\b"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["invoice_date"] = m.group(1).strip()
            break
    for pat in [r"(?:invoice\s*(?:no|number|#))[:\s#]*([A-Z0-9\-\/]+)",
                r"(?:invoice)[:\s]+([A-Z]{0,3}\d{3,10})"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["invoice_number"] = m.group(1).strip()
            break
    for pat in [r"(?:job\s*ref|work\s*order|job\s*no)[:\s#]+([A-Z0-9\-\/]+)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["job_reference"] = m.group(1).strip()
            break
    for pat in [r"(?:patient)[:\s]+([A-Za-z\s\-']{3,40}?)(?:\n|\d)",
                r"(?:client)[:\s]+([A-Z][a-z]+\s+[A-Z][a-z]+)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            name = m.group(1).strip().title()
            if len(name) > 3:
                data["patient_name"] = name
                break
    for pat in [r"(?:total\s*due|invoice\s*total|total)[:\s]+[£$]?\s*([\d,]+\.?\d*)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["invoiced_amount"] = clean_amount(m.group(1))
            break
    for pat in [r"(?:payments?\s*made|amount\s*paid)[:\s]+[£$]?\s*([\d,]+\.?\d*)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["payments_made"] = clean_amount(m.group(1))
            break
    for pat in [r"(?:balance\s*due|balance)[:\s]+[£$]?\s*([\d,]+\.?\d*)"]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            data["balance"] = clean_amount(m.group(1))
            break
    return [data]

# ─────────────────────────────────────────────
# Main processor
# ─────────────────────────────────────────────

def process_pdf(pdf_path):
    pdf_path = Path(pdf_path)
    print(f"\nProcessing: {pdf_path.name}")
    print("-" * 65)

    try:
        text, tables = extract_pdf(pdf_path)
    except Exception as e:
        print(f"  ERROR reading PDF: {e}")
        return 0

    lab = detect_lab(text)
    fmt = detect_format(text)
    print(f"  Lab: {lab or 'Unknown'}  |  Format: {fmt}")

    if fmt == "3dental":
        rows = parse_3dental(text, tables, lab)
    elif fmt == "dent8_innovate":
        rows = parse_dent8_innovate(text, tables, lab)
    elif fmt == "hall":
        rows = parse_hall(text, tables, lab)
    elif fmt == "carlkearney":
        rows = parse_carlkearney(text, tables, lab)
    elif fmt == "aestheticworld":
        rows = parse_aestheticworld(text, tables, lab)
    elif fmt == "digitalprothetics":
        rows = parse_digitalprothetics(text, tables, lab)
    elif fmt == "s4s":
        rows = parse_s4s(text, tables, lab)
    else:
        rows = parse_standard(text, lab)

    if not rows:
        print("  WARNING: No rows extracted — please check this PDF manually.")
        return 0

    for row in rows:
        ref = (row.get("job_reference") or row.get("invoice_number") or "-")[:14]
        patient = (row.get("patient_name") or "-")[:28]
        amount = f"£{row['invoiced_amount']:.2f}" if row.get("invoiced_amount") else "    -   "
        balance = f"£{row['balance']:.2f}" if row.get("balance") else "    -   "
        print(f"  • {ref:14} | {patient:28} | {amount:10} | bal: {balance}")

    count = append_rows(rows)
    print(f"  ✓ {count} row(s) logged to Invoice Tracker 2026.xlsx")
    return count

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_invoice.py <invoice.pdf or folder>")
        print("       python extract_invoice.py --dedupe   (remove duplicates from tracker)")
        sys.exit(1)

    if sys.argv[1] == "--dedupe":
        deduplicate()
        return

    target = Path(sys.argv[1])
    total = 0

    if target.is_dir():
        pdfs = sorted(set(target.rglob("*.pdf")) | set(target.rglob("*.PDF")))
        if not pdfs:
            print(f"No PDFs found in {target}")
            sys.exit(1)
        print(f"Found {len(pdfs)} PDF(s)...")
        for pdf in pdfs:
            total += process_pdf(pdf)
    elif target.is_file() and target.suffix.lower() == ".pdf":
        total += process_pdf(target)
    else:
        print(f"ERROR: '{target}' is not a valid PDF or folder.")
        sys.exit(1)

    print("\n" + "=" * 65)
    print(f"Done! {total} total rows logged.")
    print("Open Invoice Tracker 2026.xlsx to review.")

if __name__ == "__main__":
    main()
